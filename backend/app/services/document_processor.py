"""
Document processing service

Handles document loading, parsing, and text extraction
"""

from typing import List, Dict, Any
from pathlib import Path
import pypdf
from docx import Document
from langchain.text_splitter import RecursiveCharacterTextSplitter
import openpyxl
import csv

class DocumentProcessor:
    """
    Process various document formats and extract text
    """

    def __init__(self):
        """Initialize document loaders"""
        self.text_splitter = None

    async def process_file(self, file_path: str) -> Dict[str, Any]:
        """
        Process a file and extract text

        Supported formats:
        - PDF (.pdf)
        - Word (.docx, .doc)
        - Text (.txt)
        - Excel (.xlsx, .xls)
        - CSV (.csv)

        Returns:
            Dict with extracted text and metadata
        """
        path = Path(file_path)
        file_extension = path.suffix.lower()

        try:
            if file_extension == '.pdf':
                text_content = self._process_pdf(file_path)
            elif file_extension == '.docx':
                text_content = self._process_docx(file_path)
            elif file_extension == '.txt':
                text_content = self._process_txt(file_path)
            elif file_extension in ['.xlsx', '.xls']:
                text_content = self._process_excel(file_path)
            elif file_extension == '.csv':
                text_content = self._process_csv(file_path)
            else:
                raise ValueError(f"Unsupported file format: {file_extension}")

            # Combine all text
            full_text = "\n\n".join(text_content)

            return {
                "text": full_text,
                "filename": path.name,
                "file_type": file_extension,
                "size": path.stat().st_size,
                "pages": len(text_content) if file_extension == '.pdf' else None
            }
        except Exception as e:
            raise Exception(f"Error processing file {path.name}: {str(e)}")

    def _process_pdf(self, file_path: str) -> List[str]:
        """Extract text from PDF"""
        text_content = []
        with open(file_path, 'rb') as file:
            pdf_reader = pypdf.PdfReader(file)
            for page in pdf_reader.pages:
                text = page.extract_text()
                if text.strip():
                    text_content.append(text)
        return text_content

    def _process_docx(self, file_path: str) -> List[str]:
        """Extract text from Word document"""
        doc = Document(file_path)
        text_content = []
        current_text = []

        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                current_text.append(paragraph.text)

        if current_text:
            text_content.append("\n".join(current_text))

        return text_content

    def _process_txt(self, file_path: str) -> List[str]:
        """Extract text from plain text file"""
        with open(file_path, 'r', encoding='utf-8') as file:
            text = file.read()
        return [text] if text.strip() else []

    def _process_excel(self, file_path: str) -> List[str]:
        """Extract text from Excel file"""
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        text_content = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_text = [f"Sheet: {sheet_name}"]

            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    sheet_text.append(row_text)

            text_content.append("\n".join(sheet_text))

        return text_content

    def _process_csv(self, file_path: str) -> List[str]:
        """Extract text from CSV file"""
        text_content = []
        with open(file_path, 'r', encoding='utf-8') as file:
            csv_reader = csv.reader(file)
            rows = []
            for row in csv_reader:
                row_text = " | ".join(row)
                if row_text.strip():
                    rows.append(row_text)
            text_content.append("\n".join(rows))
        return text_content

    def chunk_text(self, text: str, chunk_size: int = 1000, chunk_overlap: int = 200) -> List[str]:
        """
        Split text into chunks for embedding

        Uses LangChain's RecursiveCharacterTextSplitter
        """
        if not self.text_splitter or \
           self.text_splitter._chunk_size != chunk_size or \
           self.text_splitter._chunk_overlap != chunk_overlap:
            self.text_splitter = RecursiveCharacterTextSplitter(
                chunk_size=chunk_size,
                chunk_overlap=chunk_overlap,
                length_function=len,
                separators=["\n\n", "\n", " ", ""]
            )

        chunks = self.text_splitter.split_text(text)
        return chunks

# Global document processor instance
document_processor = DocumentProcessor()
